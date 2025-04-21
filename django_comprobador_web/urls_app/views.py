from django.shortcuts import render
from .forms import UrlForm
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import os

def procesar_url(request):
    mensaje = ""
    
    if request.method == 'POST':
        form = UrlForm(request.POST)
        if form.is_valid():
            url = form.cleaned_data['url']

            # Obtener contenido de la web y extraer el título
            respuesta = requests.get(url, timeout=5)
            soup = BeautifulSoup(respuesta.text, 'html.parser')
            titulo = soup.title.string.strip() if soup.title else "Sin título"

            # Guardar en Excel
            archivo_excel = "urls_resultado.xlsx"
            if os.path.exists(archivo_excel):
                wb = load_workbook(archivo_excel)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(['URL', 'Título']) 

            ws.append([url, titulo])
            wb.save(archivo_excel)
            mensaje = f"URL procesada y guardada: {titulo}"
    else:
        form = UrlForm()

    return render(request, 'procesar_url.html', {'form': form, 'mensaje': mensaje})